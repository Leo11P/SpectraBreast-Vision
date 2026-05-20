"""Resolution sweep for the registration pipeline (single sample, many configs).

Runs the registration pipeline once for each ``(res_reg, res_pc)`` pair
declared in ``cfg.sweep.resolution_pairs`` and writes ONE Excel summary to
``<output_dir>/<sample>_sweep_resolution_summary.xlsx``.

The implementation here is a thin adapter on top of the legacy sweep logic:
the per-pair execution and the Excel layout match the original
``SpectraBreast-Registration/sweep.py`` exactly. Only the config-reading is
new (pydantic-based instead of dict-based).

If ``cfg.export.save_pointcloud`` or ``cfg.export.save_images`` is true, the
best-performing pair is re-run at the end with those flags enabled so the
final artifacts in ``output_dir`` correspond to the winning configuration.
"""

from __future__ import annotations

import os
import shutil
import time
import traceback
from pathlib import Path
from typing import Optional

import numpy as np

from ..config import RegistrationConfig
from .pipeline import load_mesh
from .pipeline_roi import run_full_pipeline_roi


_3D_ERROR_KEYS = [
    "3D_REG_bilinear_mean_mm",
    "3D_REG_bicubic_mean_mm",
    "3D_PC_bilinear_mean_mm",
    "3D_PC_bicubic_mean_mm",
]


# =============================================================================
# Stats helpers
# =============================================================================

def _stats(arr) -> tuple[float, float, int]:
    if arr is None:
        return float('nan'), float('nan'), 0
    a = np.asarray(arr, dtype=np.float64)
    if a.size == 0:
        return float('nan'), float('nan'), 0
    v = a[~np.isnan(a)]
    if v.size == 0:
        return float('nan'), float('nan'), 0
    return float(v.mean()), float(np.median(v)), int(v.size)


def _px_per_mm(data_hsi: dict, marker_side_mm: float) -> float:
    sides = []
    for corners in data_hsi.values():
        c = np.asarray(corners, dtype=np.float32)
        sides.append(float(np.mean([
            np.linalg.norm(c[(j + 1) % 4] - c[j]) for j in range(4)
        ])))
    return float(np.mean(sides)) / marker_side_mm if sides else 1.0


def _best_3d_val(row: dict) -> float:
    vals = [row.get(k, float('nan')) for k in _3D_ERROR_KEYS]
    valid = [v for v in vals if not np.isnan(v)]
    return min(valid) if valid else float('nan')


# =============================================================================
# Render cache (compute each unique resolution only once)
# =============================================================================

_RES_TOL = 1e-6


def _quantize_res(r: float) -> float:
    return round(float(r), 9)


def _collect_unique_resolutions(pairs: list[tuple[float, float]]) -> list[float]:
    seen: dict[float, float] = {}
    for r, p in pairs:
        for v in (r, p):
            k = _quantize_res(v)
            if k not in seen:
                seen[k] = v
    return sorted(seen.values(), reverse=True)


def _precompute_renders(
    mesh,
    unique_resolutions: list[float],
    margin_mm: float,
    torch_device: Optional[str],
    render_fn,
) -> dict[float, tuple]:
    if render_fn is None:
        raise RuntimeError("render_fn is None — cannot pre-compute renders.")
    cache: dict[float, tuple] = {}
    n = len(unique_resolutions)
    print(f"\n[sweep][cache] Pre-computing {n} unique renders...")
    t0_total = time.time()
    for i, res in enumerate(unique_resolutions):
        print(f"[sweep][cache] Render {i+1}/{n}: res={res} mm/px on {torch_device}")
        t0 = time.time()
        cache[_quantize_res(res)] = render_fn(
            mesh,
            resolution_mm_per_px=res,
            margin_mm=margin_mm,
            device=torch_device,
        )
        print(f"[sweep][cache]   done in {time.time() - t0:.1f}s")
    print(f"[sweep][cache] All {n} renders ready in {time.time() - t0_total:.1f}s")
    return cache


def _get_render(cache: dict[float, tuple], res: float) -> tuple:
    key = _quantize_res(res)
    if key in cache:
        return cache[key]
    for k, v in cache.items():
        if abs(k - key) < _RES_TOL:
            return v
    raise KeyError(f"Render for res={res} mm/px not in cache (keys: {sorted(cache)})")


# =============================================================================
# Single pair run
# =============================================================================

def _run_single_pair(
    cfg: RegistrationConfig,
    sample_name: str,
    aruco_dict_cv: int,
    res_reg: float,
    res_pc: float,
    torch_device: Optional[str],
    use_torch_render: bool,
    tmp_dir: str,
    render_cache: Optional[dict[float, tuple]],
    save_pointcloud_override: bool = False,
    save_images_override: bool = False,
) -> dict:
    liveview_png_path = (
        str(cfg.paths.liveview_png)
        if cfg.sweep.roi_mode and cfg.paths.liveview_png is not None
        else None
    )
    roi_mode = liveview_png_path is not None

    print(f"\n{'=' * 68}")
    print(f"  SWEEP RUN — res_reg={res_reg}  res_pc={res_pc}  ROI={'YES' if roi_mode else 'NO'}")
    print(f"{'=' * 68}")
    t0 = time.time()
    os.makedirs(tmp_dir, exist_ok=True)
    dual = abs(res_reg - res_pc) > 1e-6

    precomputed_render = None
    precomputed_render_reg = None
    if use_torch_render and render_cache is not None:
        precomputed_render = _get_render(render_cache, res_pc)
        precomputed_render_reg = _get_render(render_cache, res_reg) if dual else precomputed_render

    result = run_full_pipeline_roi(
        hsi_hdr_path             = str(cfg.paths.hsi_hdr),
        mesh_path                = str(cfg.paths.mesh),
        aruco_json_path          = str(cfg.paths.aruco_json),
        output_dir               = tmp_dir,
        liveview_png_path        = liveview_png_path,
        roi_align_cfg            = cfg.roi.model_dump() if roi_mode else None,
        hsi_extraction_method    = cfg.detection.hsi_extraction_method,
        aruco_dict_type          = aruco_dict_cv,
        suspicious_pixels_hsi    = None,
        render_resolution_mm     = res_pc,
        render_resolution_reg_mm = res_reg,
        render_margin_mm         = cfg.render.margin_mm,
        marker_side_mm           = cfg.detection.marker_side_mm,
        use_subpix               = cfg.detection.use_subpix,
        subpix_winsize           = cfg.detection.subpix_winsize,
        border_px                = cfg.pointcloud.border_px,
        reflectance_norm         = cfg.pointcloud.reflectance_norm,
        pc_chunk_size            = cfg.pointcloud.pc_chunk_size,
        export_ply_file          = cfg.export.ply  if save_pointcloud_override else False,
        export_npy_file          = cfg.export.npy  if save_pointcloud_override else False,
        export_csv_file          = cfg.export.csv  if save_pointcloud_override else False,
        csv_max_points           = cfg.export.csv_max_points,
        save_pointcloud          = save_pointcloud_override,
        save_images              = save_images_override,
        sample_name              = sample_name,
        precomputed_render       = precomputed_render,
        precomputed_render_reg   = precomputed_render_reg,
    )

    # Extract metrics
    err2d_px = result.get("err2d_px")
    err3d_dict = result.get("err3d_dict")
    err3d_dict_pc = result.get("err3d_dict_pc")
    data_hsi = result.get("data_hsi")
    roi_info = result.get("roi_align_info")

    pxpermm = _px_per_mm(data_hsi, cfg.detection.marker_side_mm) if data_hsi else float("nan")
    if err2d_px is not None:
        e = np.asarray(err2d_px, dtype=np.float64)
        m2px, med2px, _ = _stats(e)
        if not np.isnan(pxpermm) and pxpermm > 0:
            m2mm, med2mm, _ = _stats(e / pxpermm)
        else:
            m2mm = med2mm = float("nan")
    else:
        m2px = med2px = m2mm = med2mm = float("nan")

    if err3d_dict is not None:
        m_rb_bil, md_rb_bil, _ = _stats(err3d_dict.get("bilinear"))
        m_rb_bic, md_rb_bic, _ = _stats(err3d_dict.get("bicubic"))
    else:
        m_rb_bil = md_rb_bil = m_rb_bic = md_rb_bic = float("nan")

    if err3d_dict_pc is not None:
        m_pc_bil, md_pc_bil, _ = _stats(err3d_dict_pc.get("bilinear"))
        m_pc_bic, md_pc_bic, _ = _stats(err3d_dict_pc.get("bicubic"))
    else:
        m_pc_bil, md_pc_bil = m_rb_bil, md_rb_bil
        m_pc_bic, md_pc_bic = m_rb_bic, md_rb_bic

    return {
        "res_reg_mm_pix":            res_reg,
        "res_pc_mm_pix":             res_pc,
        "roi_mode":                  "yes" if roi_mode else "no",
        "2D_mean_px": m2px, "2D_median_px": med2px,
        "2D_mean_mm": m2mm, "2D_median_mm": med2mm,
        "3D_REG_bilinear_mean_mm":   m_rb_bil, "3D_REG_bilinear_median_mm": md_rb_bil,
        "3D_REG_bicubic_mean_mm":    m_rb_bic, "3D_REG_bicubic_median_mm":  md_rb_bic,
        "3D_PC_bilinear_mean_mm":    m_pc_bil, "3D_PC_bilinear_median_mm":  md_pc_bil,
        "3D_PC_bicubic_mean_mm":     m_pc_bic, "3D_PC_bicubic_median_mm":   md_pc_bic,
        "roi_inliers":   roi_info["n_inliers"]            if roi_info else 0,
        "roi_matches":   roi_info["n_good_matches"]       if roi_info else 0,
        "roi_reproj_px": roi_info["reproj_error_mean_px"] if roi_info else float("nan"),
        "elapsed_s":     round(time.time() - t0, 2),
        "status":        "ok",
    }


def _empty_row(res_reg: float, res_pc: float, status: str, roi_mode: bool = False) -> dict:
    nan = float("nan")
    return {
        "res_reg_mm_pix": res_reg, "res_pc_mm_pix": res_pc,
        "roi_mode": "yes" if roi_mode else "no",
        "2D_mean_px": nan, "2D_median_px": nan, "2D_mean_mm": nan, "2D_median_mm": nan,
        "3D_REG_bilinear_mean_mm":   nan, "3D_REG_bilinear_median_mm": nan,
        "3D_REG_bicubic_mean_mm":    nan, "3D_REG_bicubic_median_mm":  nan,
        "3D_PC_bilinear_mean_mm":    nan, "3D_PC_bilinear_median_mm":  nan,
        "3D_PC_bicubic_mean_mm":     nan, "3D_PC_bicubic_median_mm":   nan,
        "roi_inliers": 0, "roi_matches": 0, "roi_reproj_px": nan,
        "elapsed_s": 0.0, "status": status,
    }


def _find_best_row(rows: list[dict]) -> tuple[int, float]:
    best_idx, best_val = -1, float("inf")
    for i, row in enumerate(rows):
        if row.get("status") != "ok":
            continue
        v = _best_3d_val(row)
        if not np.isnan(v) and v < best_val:
            best_val, best_idx = v, i
    return best_idx, (best_val if best_idx >= 0 else float("nan"))


# =============================================================================
# Excel
# =============================================================================

def _write_summary_xlsx(rows: list[dict], output_path: str, sample_name: str,
                         best_row_idx: int = -1) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sweep Summary"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="305496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")
    norm_font = Font(name="Calibri", size=10)
    best_font = Font(bold=True, color="FFFFFF", size=10)
    best_fill = PatternFill("solid", start_color="00B050")

    ws.cell(row=1, column=1, value=f"Resolution Sweep — {sample_name}").font = \
        Font(bold=True, size=14, color="1F4E78")
    ws.cell(row=2, column=1,
            value=f"{len(rows)} configurations | {time.strftime('%Y-%m-%d %H:%M:%S')}").font = \
        Font(italic=True, color="595959")
    if best_row_idx >= 0:
        br = rows[best_row_idx]
        bv = _best_3d_val(br)
        ws.cell(row=3, column=1,
                value=f"★ Best: res_reg={br['res_reg_mm_pix']} res_pc={br['res_pc_mm_pix']} "
                      f"→ 3D error min = {bv:.4f} mm (row {best_row_idx + 1})").font = \
            Font(italic=True, bold=True, color="00B050", size=11)

    columns = [
        ("res_reg_mm_pix",            "res_reg\n(mm/px)"),
        ("res_pc_mm_pix",             "res_pc\n(mm/px)"),
        ("roi_mode",                  "ROI"),
        ("2D_mean_px",                "2D mean\n(px)"),
        ("2D_median_px",              "2D median\n(px)"),
        ("2D_mean_mm",                "2D mean\n(mm)"),
        ("2D_median_mm",              "2D median\n(mm)"),
        ("3D_REG_bilinear_mean_mm",   "3D REG bil\nmean (mm)"),
        ("3D_REG_bilinear_median_mm", "3D REG bil\nmedian (mm)"),
        ("3D_REG_bicubic_mean_mm",    "3D REG bic\nmean (mm)"),
        ("3D_REG_bicubic_median_mm",  "3D REG bic\nmedian (mm)"),
        ("3D_PC_bilinear_mean_mm",    "3D PC bil\nmean (mm)"),
        ("3D_PC_bilinear_median_mm",  "3D PC bil\nmedian (mm)"),
        ("3D_PC_bicubic_mean_mm",     "3D PC bic\nmean (mm)"),
        ("3D_PC_bicubic_median_mm",   "3D PC bic\nmedian (mm)"),
        ("roi_inliers",               "ROI inliers"),
        ("roi_matches",               "ROI matches"),
        ("roi_reproj_px",             "ROI reproj (px)"),
        ("elapsed_s",                 "Elapsed (s)"),
        ("status",                    "Status"),
    ]

    HEADER_ROW = 5
    for ci, (_, label) in enumerate(columns, start=1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = border
    ws.row_dimensions[HEADER_ROW].height = 36

    def _fmt(v):
        return "N/A" if isinstance(v, float) and np.isnan(v) else v

    for row_i, row in enumerate(rows, start=HEADER_ROW + 1):
        is_best = (row_i - HEADER_ROW - 1) == best_row_idx
        for ci, (key, _) in enumerate(columns, start=1):
            val = row.get(key, "")
            c = ws.cell(row=row_i, column=ci, value=_fmt(val))
            c.alignment = center; c.border = border
            if is_best:
                c.font = best_font; c.fill = best_fill
            else:
                c.font = norm_font
            if isinstance(val, float) and not np.isnan(val):
                if key.endswith(("_mm", "_px", "_s")):
                    c.number_format = "0.0000"

    for ci, (_, label) in enumerate(columns, start=1):
        max_len = max(len(line) for line in label.split("\n"))
        ws.column_dimensions[get_column_letter(ci)].width = max(max_len + 3, 12)
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=4)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"\n[sweep] Excel summary saved → {output_path}")


# =============================================================================
# Entry point
# =============================================================================

def run_sweep(
    cfg: RegistrationConfig,
    aruco_dict_cv: int,
    torch_device: Optional[str],
    use_torch_render: bool,
    sample_name: str,
    render_fn=None,
) -> str:
    """Run the sweep on a single sample and return the Excel summary path."""
    pairs = [(float(r), float(p)) for r, p in cfg.sweep.resolution_pairs]
    roi_mode = bool(cfg.sweep.roi_mode)

    if roi_mode and cfg.paths.liveview_png is None:
        raise ValueError("sweep.roi_mode=true requires registration.paths.liveview_png.")

    save_pointcloud_cfg = cfg.export.save_pointcloud
    save_images_cfg = cfg.export.save_images
    do_best_rerun = save_pointcloud_cfg or save_images_cfg

    print("\n" + "=" * 68)
    print(f"  SWEEP — {len(pairs)} (reg, pc) pairs   ROI={'YES' if roi_mode else 'NO'}")
    print("=" * 68)
    for i, (r, p) in enumerate(pairs):
        print(f"  [{i + 1:2d}/{len(pairs)}]  res_reg={r:>6.3f}  res_pc={p:>6.3f}")
    if do_best_rerun:
        print(f"\n  [INFO] Best pair will be re-run with "
              f"save_pointcloud={save_pointcloud_cfg} save_images={save_images_cfg}")

    base_out = str(cfg.paths.output_dir)
    os.makedirs(base_out, exist_ok=True)
    sweep_tmp = os.path.join(base_out, "_sweep_tmp")
    os.makedirs(sweep_tmp, exist_ok=True)
    print(f"[sweep] Temp subdirs in: {sweep_tmp}")

    # Pre-compute renders if GPU available
    render_cache: Optional[dict[float, tuple]] = None
    if use_torch_render and render_fn is not None:
        mesh = load_mesh(str(cfg.paths.mesh), scale_m_to_mm=True)
        unique_res = _collect_unique_resolutions(pairs)
        print(f"\n[sweep] Unique resolutions: {[str(r) for r in unique_res]}")
        render_cache = _precompute_renders(
            mesh=mesh,
            unique_resolutions=unique_res,
            margin_mm=cfg.render.margin_mm,
            torch_device=torch_device,
            render_fn=render_fn,
        )
    else:
        print("[sweep] WARNING: render_fn unavailable — each run re-renders internally.")

    # Main loop
    rows: list[dict] = []
    for i, (res_reg, res_pc) in enumerate(pairs):
        tag = f"reg{str(res_reg).replace('.', 'p')}_pc{str(res_pc).replace('.', 'p')}"
        tmp_dir = os.path.join(sweep_tmp, f"run{i+1:02d}_{tag}")
        try:
            row = _run_single_pair(
                cfg=cfg, sample_name=sample_name, aruco_dict_cv=aruco_dict_cv,
                res_reg=res_reg, res_pc=res_pc,
                torch_device=torch_device, use_torch_render=use_torch_render,
                tmp_dir=tmp_dir, render_cache=render_cache,
                save_pointcloud_override=False, save_images_override=False,
            )
        except Exception as e:
            print(f"\n[sweep] ERROR on ({res_reg}, {res_pc}): {e}")
            traceback.print_exc()
            row = _empty_row(res_reg, res_pc,
                             status=f"error: {type(e).__name__}: {e}",
                             roi_mode=roi_mode)
        rows.append(row)
        print(f"\n[sweep] Progress: {i + 1}/{len(pairs)} done.\n")

    best_idx, best_val = _find_best_row(rows)
    if best_idx >= 0:
        br = rows[best_idx]
        print(f"\n[sweep] ★ Best pair (row {best_idx + 1}): "
              f"res_reg={br['res_reg_mm_pix']}  res_pc={br['res_pc_mm_pix']}  "
              f"→ 3D error min = {best_val:.4f} mm")

        if do_best_rerun:
            print(f"\n[sweep] Re-running best pair with save flags enabled...")
            try:
                _run_single_pair(
                    cfg=cfg, sample_name=sample_name, aruco_dict_cv=aruco_dict_cv,
                    res_reg=br["res_reg_mm_pix"], res_pc=br["res_pc_mm_pix"],
                    torch_device=torch_device, use_torch_render=use_torch_render,
                    tmp_dir=base_out, render_cache=render_cache,
                    save_pointcloud_override=save_pointcloud_cfg,
                    save_images_override=save_images_cfg,
                )
            except Exception as e:
                print(f"[sweep] ERROR re-running best pair: {e}")
                traceback.print_exc()

    try:
        shutil.rmtree(sweep_tmp)
        print(f"[sweep] Removed temp dir: {sweep_tmp}")
    except Exception as e:
        print(f"[sweep] Could not remove {sweep_tmp}: {e}")

    out_path = os.path.join(base_out, f"{sample_name}_sweep_resolution_summary.xlsx")
    _write_summary_xlsx(rows, out_path, sample_name=sample_name, best_row_idx=best_idx)

    n_ok = sum(1 for r in rows if r["status"] == "ok")
    print("\n" + "=" * 68)
    print("SWEEP COMPLETED")
    print("=" * 68)
    print(f"  Pairs run    : {len(rows)}")
    print(f"  OK / Failed  : {n_ok} / {len(rows) - n_ok}")
    print(f"  Excel summary: {out_path}")
    print("=" * 68)
    return out_path


__all__ = ["run_sweep"]
