"""Mode dispatcher for the registration pipeline.

This module is the single entry point used by ``spectra registration`` (and by
``spectra full`` via :mod:`spectra.orchestrator`). It reads a
:class:`~spectra.config.RegistrationConfig` and runs the right backend:

    mode="single" → :func:`spectra.registration.pipeline_roi.run_full_pipeline_roi`
                    with ``liveview_png_path=None`` (delegates to
                    ``run_full_pipeline`` internally — see pipeline_roi.py)
    mode="roi"    → :func:`spectra.registration.pipeline_roi.run_full_pipeline_roi`
                    with the LiveView PNG enabled
    mode="batch"  → :func:`spectra.registration.batch.run_batch`
    mode="sweep"  → :func:`spectra.registration.sweep.run_sweep`

The dispatcher does NOT auto-derive paths: the orchestrator
(:mod:`spectra.orchestrator`) is responsible for filling
``cfg.paths.{hsi_hdr,mesh,aruco_json,liveview_png,output_dir}`` from
``sample_name`` + ``data_root`` + ``results_root`` BEFORE calling this. When
:func:`run_registration` is invoked directly with missing paths it raises a
clean error.

Behavior is otherwise identical to the legacy ``main.py``.
"""

from __future__ import annotations

import os
import sys
import time
from pathlib import Path
from typing import Any, Optional

import cv2

from ..config import RegistrationConfig

import json
from typing import Dict


def read_vision_timings(vision_dir: Path) -> Optional[Dict[str, Any]]:
    """Rilegge <vision_dir>/timings.json. None se assente/illeggibile.

    vision_dir = RESULTS/<sample>/vision
    """
    path = Path(vision_dir) / "timings.json"
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None
    return data if isinstance(data, dict) else None


def _vision_dir_for(results_root, sample_name) -> Path:
    return Path(results_root) / str(sample_name) / "vision"

# Render GPU is optional (torch / cupy may be absent).
try:
    import torch
    from .render_gpu import render_orthographic_topview_gpu
    _TORCH_AVAILABLE = True
    _CUDA_AVAILABLE = torch.cuda.is_available()
except Exception:
    render_orthographic_topview_gpu = None
    torch = None  # type: ignore[assignment]
    _TORCH_AVAILABLE = False
    _CUDA_AVAILABLE = False


_ARUCO_DICT_MAP = {
    "4X4_50":  cv2.aruco.DICT_4X4_50,
    "4X4_100": cv2.aruco.DICT_4X4_100,
    "5X5_50":  cv2.aruco.DICT_5X5_50,
    "5X5_100": cv2.aruco.DICT_5X5_100,
    "6X6_50":  cv2.aruco.DICT_6X6_50,
    "6X6_100": cv2.aruco.DICT_6X6_100,
}


# =============================================================================
# Path validation
# =============================================================================

def _missing_paths(cfg: RegistrationConfig) -> list[str]:
    """Return human-readable names of registration paths that are still ``None``."""
    p = cfg.paths
    missing: list[str] = []
    # Always required:
    if p.hsi_hdr is None:
        missing.append("paths.hsi_hdr")
    if cfg.mode != "batch":
        # batch discovers per-sample paths itself
        if p.mesh is None:
            missing.append("paths.mesh")
        if p.aruco_json is None:
            missing.append("paths.aruco_json")
    if p.output_dir is None:
        missing.append("paths.output_dir")
    needs_liveview = (
        cfg.mode == "roi"
        or (cfg.mode == "sweep" and cfg.sweep.roi_mode)
    )
    if needs_liveview and p.liveview_png is None:
        missing.append("paths.liveview_png")
    return missing


def _validate_inputs_exist(cfg: RegistrationConfig) -> list[str]:
    """Return human-readable error messages for missing files on disk."""
    p = cfg.paths
    errors: list[str] = []
    checks = [(p.hsi_hdr, "HSI .hdr")]
    if cfg.mode in ("single", "roi", "sweep"):
        checks += [(p.mesh, "Mesh"), (p.aruco_json, "ArUco JSON")]
    if cfg.mode == "roi" or (cfg.mode == "sweep" and cfg.sweep.roi_mode):
        checks += [(p.liveview_png, "LiveView PNG")]
    for path, label in checks:
        if path is not None and not Path(path).exists():
            errors.append(f"{label} not found: {path}")
    return errors


# =============================================================================
# Helpers shared by single/roi runs
# =============================================================================

def _device_string(force_cpu: bool = False) -> Optional[str]:
    if not _TORCH_AVAILABLE:
        return None
    if force_cpu:
        return "cpu"
    return "cuda" if _CUDA_AVAILABLE else "cpu"


def _print_summary(cfg: RegistrationConfig, sample_name: str) -> None:
    print("\n" + "=" * 68)
    print("SPECTRA REGISTRATION — SUMMARY")
    print("=" * 68)
    print(f"  Mode          : {cfg.mode.upper()}")
    print(f"  Sample        : {sample_name}")
    print(f"  HSI           : {cfg.paths.hsi_hdr}")
    if cfg.mode != "batch":
        print(f"  Mesh          : {cfg.paths.mesh}")
        print(f"  ArUco JSON    : {cfg.paths.aruco_json}")
        if cfg.mode == "roi" or (cfg.mode == "sweep" and cfg.sweep.roi_mode):
            print(f"  LiveView PNG  : {cfg.paths.liveview_png}")
    print(f"  Marker side   : {cfg.detection.marker_side_mm} mm")

    if cfg.mode == "sweep":
        print(f"  Render        : sweep over {len(cfg.sweep.resolution_pairs)} pairs")
        print(f"  ROI sweep     : {'YES' if cfg.sweep.roi_mode else 'NO'}")
    else:
        print(f"  Render        : reg={cfg.render.resolution_reg_mm_per_px} "
              f"pc={cfg.render.resolution_mm_per_px} mm/px")

    print(f"  Save PC       : {'YES' if cfg.export.save_pointcloud else 'NO'}")
    print(f"  Save images   : {'YES' if cfg.export.save_images else 'NO'}")
    print(f"  Output        : {cfg.paths.output_dir}")
    print("=" * 68)


# =============================================================================
# Dispatcher
# =============================================================================

def run_registration(
    cfg: RegistrationConfig,
    *,
    sample_name: str = "sample",
    data_root: Optional[Path] = None,
    results_root: Optional[Path] = None,
    force_cpu: bool = False,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Run the registration stage in the mode declared by ``cfg.mode``.

    Parameters
    ----------
    cfg :
        A fully-resolved ``RegistrationConfig`` (paths must already be filled in
        by the caller — the orchestrator does that automatically for
        ``spectra full``).
    sample_name :
        Used as the filename prefix for outputs (Excel report, point cloud,
        registration PNGs).
    data_root, results_root :
        Required only for ``mode == "batch"`` (the batch discoverer scans
        ``data_root`` and writes per-sample outputs under
        ``results_root/<sample>/registration``).
    force_cpu :
        If true, disables GPU raycasting even if torch/cupy are available.
    dry_run :
        If true, validates paths and prints the summary but does not run.

    Returns
    -------
    dict
        Backend-specific result dictionary. For batch/sweep, contains the
        Excel summary path under ``"summary_xlsx"``.
    """
    missing = _missing_paths(cfg)
    if missing:
        raise ValueError(
            "Registration config is missing required paths "
            f"(none of these can be null when running standalone): {missing}. "
            "Either set them explicitly under `registration.paths:` in the YAML, "
            "or run via `spectra full` so the orchestrator derives them from "
            "sample_name + data_root + results_root."
        )

    errors = _validate_inputs_exist(cfg)
    if errors:
        for e in errors:
            print(f"[Validate] MISSING — {e}")
        raise FileNotFoundError(
            f"Registration cannot start: {len(errors)} input file(s) missing."
        )

    _print_summary(cfg, sample_name)

    if dry_run:
        print("\n[Dry-run] Configuration valid; pipeline NOT executed.")
        return {"dry_run": True}

    aruco_dict_cv = _ARUCO_DICT_MAP.get(cfg.detection.aruco_dict.upper(), cv2.aruco.DICT_4X4_50)
    device_str = _device_string(force_cpu=force_cpu)
    use_gpu_render = _TORCH_AVAILABLE and (render_orthographic_topview_gpu is not None)

    os.makedirs(cfg.paths.output_dir, exist_ok=True)  # type: ignore[arg-type]

    # ----- mode == "batch" --------------------------------------------------
    if cfg.mode == "batch":
        if data_root is None or results_root is None:
            raise ValueError(
                "Batch mode requires data_root and results_root to be passed. "
                "These are normally provided by the unified CLI / orchestrator."
            )
        from .batch import run_batch
        xlsx_path = run_batch(
            cfg=cfg,
            aruco_dict_cv=aruco_dict_cv,
            torch_device=device_str,
            use_torch_render=use_gpu_render,
            data_root=Path(data_root),
            results_root=Path(results_root),
            render_fn=render_orthographic_topview_gpu if use_gpu_render else None,
        )
        return {"summary_xlsx": xlsx_path}

    # ----- mode == "sweep" --------------------------------------------------
    if cfg.mode == "sweep":
        from .sweep import run_sweep
        xlsx_path = run_sweep(
            cfg=cfg,
            aruco_dict_cv=aruco_dict_cv,
            torch_device=device_str,
            use_torch_render=use_gpu_render,
            sample_name=sample_name,
            render_fn=render_orthographic_topview_gpu if use_gpu_render else None,
        )
        return {"summary_xlsx": xlsx_path}

    # ----- mode == "single" / "roi" -----------------------------------------
    return _run_single_or_roi(
        cfg=cfg,
        sample_name=sample_name,
        aruco_dict_cv=aruco_dict_cv,
        device_str=device_str,
        use_gpu_render=use_gpu_render,
        results_root=results_root,
    )


def _run_single_or_roi(
    cfg: RegistrationConfig,
    sample_name: str,
    aruco_dict_cv: int,
    device_str: Optional[str],
    use_gpu_render: bool,
    results_root: Optional[Path] = None,
) -> dict[str, Any]:
    """Single / ROI path — mirrors the second half of legacy main.py."""
    from .pipeline import load_mesh, save_render, save_turbo_render
    from .pipeline_roi import run_full_pipeline_roi

    res_pc = cfg.render.resolution_mm_per_px
    res_reg = cfg.render.resolution_reg_mm_per_px
    dual = abs(res_reg - res_pc) > 1e-6

    output_dir = str(cfg.paths.output_dir)
    save_pointcloud = cfg.export.save_pointcloud
    save_images = cfg.export.save_images

    precomputed_render = None
    precomputed_render_reg = None

    _t_render_total = 0.0   # somma render pc (+ reg se dual)

    # Pre-render on GPU once, so single+roi pipelines reuse them.
    if use_gpu_render and render_orthographic_topview_gpu is not None:
        mesh = load_mesh(str(cfg.paths.mesh), scale_m_to_mm=True)

        print(f"\n[Render PC] {res_pc} mm/px on {device_str}...")
        t0 = time.time()
        precomputed_render = render_orthographic_topview_gpu(
            mesh,
            resolution_mm_per_px=res_pc,
            margin_mm=cfg.render.margin_mm,
            device=device_str,
        )
        _dt = time.time() - t0
        _t_render_total += _dt
        print(f"[Render PC] Completed in {_dt:.1f}s")
        if save_images:
            r_rgb, d_map, xyz_map, _, _ = precomputed_render
            save_render(r_rgb, d_map, output_dir=output_dir,
                        prefix=f"{sample_name}_render_pc")
            save_turbo_render(
                xyz_map,
                str(Path(output_dir) / f"{sample_name}_render_pc_turbo.png"),
            )

        if dual:
            print(f"\n[Render REG] {res_reg} mm/px on {device_str}...")
            t0 = time.time()
            precomputed_render_reg = render_orthographic_topview_gpu(
                mesh,
                resolution_mm_per_px=res_reg,
                margin_mm=cfg.render.margin_mm,
                device=device_str,
            )
            _dt = time.time() - t0
            _t_render_total += _dt
            print(f"[Render REG] Completed in {_dt:.1f}s")
            if save_images:
                r_rgb, d_map, xyz_map, _, _ = precomputed_render_reg
                save_render(r_rgb, d_map, output_dir=output_dir,
                            prefix=f"{sample_name}_render_reg")
                save_turbo_render(
                    xyz_map,
                    str(Path(output_dir) / f"{sample_name}_render_reg_turbo.png"),
                )
        else:
            precomputed_render_reg = precomputed_render

    # ROI-only params
    liveview_png_path = str(cfg.paths.liveview_png) if cfg.mode == "roi" else None
    roi_align_cfg = cfg.roi.model_dump() if cfg.mode == "roi" else None

    t0 = time.time()
    result = run_full_pipeline_roi(
        hsi_hdr_path             = str(cfg.paths.hsi_hdr),
        mesh_path                = str(cfg.paths.mesh),
        aruco_json_path          = str(cfg.paths.aruco_json),
        output_dir               = output_dir,
        liveview_png_path        = liveview_png_path,
        roi_align_cfg            = roi_align_cfg,
        hsi_extraction_method    = cfg.detection.hsi_extraction_method,
        aruco_dict_type          = aruco_dict_cv,
        suspicious_pixels_hsi    = None,
        render_resolution_mm     = res_pc,
        render_resolution_reg_mm = res_reg,
        render_margin_mm         = cfg.render.margin_mm,
        marker_side_mm           = cfg.detection.marker_side_mm,
        use_subpix               = cfg.detection.use_subpix,
        subpix_winsize           = cfg.detection.subpix_winsize,
        equalize_method          = cfg.detection.equalize_method,
        tune_detector            = cfg.detection.tune_detector,
        border_px                = cfg.pointcloud.border_px,
        reflectance_norm         = cfg.pointcloud.reflectance_norm,
        pc_chunk_size            = cfg.pointcloud.pc_chunk_size,
        export_ply_file          = cfg.export.ply  if save_pointcloud else False,
        export_npy_file          = cfg.export.npy  if save_pointcloud else False,
        export_csv_file          = cfg.export.csv  if save_pointcloud else False,
        csv_max_points           = cfg.export.csv_max_points,
        save_pointcloud          = save_pointcloud,
        save_images              = save_images,
        sample_name              = sample_name,
        precomputed_render       = precomputed_render,
        precomputed_render_reg   = precomputed_render_reg,
    )
    elapsed = time.time() - t0          # = registration + export (timer post-render)
    _reg_total = _t_render_total + elapsed   # render GPU + pipeline/export

    print("\n" + "=" * 68)
    print("REGISTRATION COMPLETED")
    print("=" * 68)
    print(f"  Total time     : {elapsed:.1f}s  ({elapsed/60:.1f} min)")
    n_pts = result.get("n_valid", 0)
    n_bands = result.get("bands", 0)
    ...
    print(f"  Output         : {output_dir}")
    print("=" * 68)

    # --- Timings nella sheet Excel ---------------------------------------
    excel_path = str(Path(output_dir) / f"{sample_name}_registration_errors.xlsx")
    timings = {
        "render_total_s": round(_t_render_total, 4),
        "pipeline_s": round(elapsed, 4),
        "reg_total_s": round(_reg_total, 4),
    }
    result["_timings"] = timings   # esposto per l'orchestrator (totale vision+reg)
    try:
        _append_single_timings_sheet(
            excel_path=excel_path,
            sample_name=sample_name,
            timings=timings,
            vision_dir = _vision_dir_for(results_root, sample_name) if results_root else None
        )
    except Exception as exc:
        print(f"[yellow]Could not append Timings sheet: {exc}[/yellow]")

    return result




def _append_single_timings_sheet(
    *,
    excel_path: str,
    sample_name: str,
    timings: Dict[str, Any],
    vision_dir: Optional[Path],
) -> None:
    """Apre l'Excel di registration e aggiunge/sostituisce la sheet 'Timings'."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not Path(excel_path).is_file():
        print(f"[yellow]Timings: Excel not found ({excel_path}); skip.[/yellow]")
        return

    vt = read_vision_timings(vision_dir) if vision_dir else None
    vis_total = vt.get("total_seconds") if vt else None
    grand_total = (vis_total + timings["reg_total_s"]) if vis_total is not None else None

    wb = load_workbook(excel_path)
    if "Timings" in wb.sheetnames:
        del wb["Timings"]
    ws = wb.create_sheet("Timings")

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", start_color="305496")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    norm = Font(name="Calibri", size=10); bold = Font(name="Calibri", size=10, bold=True)
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    vis_fill = PatternFill("solid", start_color="E2EFDA")
    reg_fill = PatternFill("solid", start_color="D9E1F2")
    tot_fill = PatternFill("solid", start_color="FFE699")

    ws.cell(row=1, column=1, value="Timings — mode: SINGLE").font = \
        Font(bold=True, size=14, color="1F4E78")

    ri = [3]  # mutabile per closure
    def section(name):
        c1 = ws.cell(row=ri[0], column=1, value=name); c1.font = hdr_font; c1.fill = hdr_fill; c1.border = border
        c2 = ws.cell(row=ri[0], column=2, value="seconds"); c2.font = hdr_font; c2.fill = hdr_fill; c2.border = border
        ri[0] += 1
    def kv(label, val, fill, b=False):
        c1 = ws.cell(row=ri[0], column=1, value=label); c1.font = bold if b else norm; c1.alignment = left; c1.border = border; c1.fill = fill
        c2 = ws.cell(row=ri[0], column=2, value=("N/A" if val is None else val)); c2.font = bold if b else norm; c2.alignment = center; c2.border = border; c2.fill = fill
        if isinstance(val, (int, float)): c2.number_format = "0.0000"
        ri[0] += 1

    section("GLOBAL TOTALS")
    kv("Total (code start → end)", round(grand_total, 4) if grand_total is not None else None, tot_fill, True)
    kv("Vision total (launch → registration start)", vis_total, vis_fill, True)
    kv("Registration total (render + pipeline)", timings["reg_total_s"], reg_fill, True)
    ri[0] += 1

    section("VISION SUB-PHASES")
    if vt and vt.get("phases"):
        for p in vt["phases"]:
            kv(p["label"], p["seconds"], vis_fill)
    else:
        kv("Vision phases", None, vis_fill)
    ri[0] += 1

    section("REGISTRATION")
    kv("GPU render total (pc + reg)", timings["render_total_s"], reg_fill)
    kv("Registration + export", timings["pipeline_s"], reg_fill)
    kv("Registration total", timings["reg_total_s"], reg_fill, True)

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18
    wb.save(excel_path)
    print(f"[dim]Timings sheet added → {excel_path}[/dim]")





__all__ = ["run_registration"]