"""Batch registration over multiple samples under the new DATA/<sample>/ layout.

Discovery rule
==============
For every subdirectory ``<data_root>/<sample>/`` that satisfies ALL of:

1. It contains ``input_registration/<sample>_raw.hdr`` (the HSI cube).
2. ``<results_root>/<sample>/vision/surface_mesh.ply`` exists.
3. ``<results_root>/<sample>/vision/aruco_markers_3d.json`` exists.
4. (When ``cfg.batch.roi_mode`` is true) ``input_registration/<sample>_raw.png``
   also exists.
5. If ``cfg.batch.sample_regex`` is set, ``<sample>`` matches it.

…the batch runs the per-sample registration pipeline and writes outputs to
``<results_root>/<sample>/registration/``. After every sample, an aggregate
Excel is written to ``<results_root>/_batch/batch_summary.xlsx``.

The structure of the Excel (columns, conditional formatting) is identical to
the legacy ``SpectraBreast-Registration/batch.py`` so existing analysis
scripts keep working.
"""

from __future__ import annotations

import os
import re
import time
import traceback
from pathlib import Path
from typing import Any, Optional

import numpy as np

from ..config import RegistrationConfig
from .pipeline import load_mesh
from .pipeline_roi import run_full_pipeline_roi


# =============================================================================
# Stats helpers (lifted from the legacy batch.py — unchanged)
# =============================================================================

def _stats(arr) -> tuple[float, float, int]:
    if arr is None:
        return float('nan'), float('nan'), 0
    a = np.asarray(arr, dtype=np.float64)
    if a.size == 0:
        return float('nan'), float('nan'), 0
    v = a[~np.isnan(a)]
    if v.size == 0:
        return float('nan'), float('nan'), 0
    return float(v.mean()), float(np.median(v)), int(v.size)


def _px_per_mm_from_data(data_hsi: dict, marker_side_mm: float) -> float:
    sides = []
    for corners in data_hsi.values():
        c = np.asarray(corners, dtype=np.float32)
        sides.append(float(np.mean([
            np.linalg.norm(c[(j + 1) % 4] - c[j]) for j in range(4)
        ])))
    return float(np.mean(sides)) / marker_side_mm if sides else 1.0


# =============================================================================
# Discovery
# =============================================================================

def discover_samples(
    data_root: Path,
    results_root: Path,
    roi_mode: bool,
    sample_regex: Optional[str] = None,
) -> list[dict]:
    """Walk ``data_root`` and return one dict per ready-to-run sample."""
    data_root = Path(data_root)
    results_root = Path(results_root)
    if not data_root.is_dir():
        raise FileNotFoundError(f"[batch] data_root not found: {data_root}")

    pattern = re.compile(sample_regex) if sample_regex else None
    pairs: list[dict] = []
    skipped: list[tuple[str, str]] = []

    for entry in sorted(data_root.iterdir()):
        if not entry.is_dir():
            continue
        sample = entry.name
        if pattern is not None and not pattern.search(sample):
            continue

        hdr = entry / "input_registration" / f"{sample}_raw.hdr"
        png = entry / "input_registration" / f"{sample}_raw.png"
        mesh = results_root / sample / "vision" / "surface_mesh.ply"
        aruco = results_root / sample / "vision" / "aruco_markers_3d.json"

        missing = []
        if not hdr.is_file():
            missing.append(f"input_registration/{sample}_raw.hdr")
        if not mesh.is_file():
            missing.append("vision/surface_mesh.ply")
        if not aruco.is_file():
            missing.append("vision/aruco_markers_3d.json")
        if roi_mode and not png.is_file():
            missing.append(f"input_registration/{sample}_raw.png")
        if missing:
            skipped.append((sample, "missing: " + ", ".join(missing)))
            continue

        pairs.append({
            "sample_name":       sample,
            "hsi_hdr_path":      str(hdr),
            "mesh_path":         str(mesh),
            "aruco_path":        str(aruco),
            "liveview_png_path": str(png) if roi_mode else None,
            "output_dir":        str(results_root / sample / "registration"),
        })

    print(f"\n[batch] Scanned: {data_root}")
    print(f"[batch] ROI mode: {'YES' if roi_mode else 'NO'}")
    if sample_regex:
        print(f"[batch] Regex filter: {sample_regex!r}")
    print(f"[batch] Ready samples: {len(pairs)}")
    for p in pairs:
        print(f"  + {p['sample_name']}")
        print(f"      hsi  : {Path(p['hsi_hdr_path']).relative_to(data_root)}")
        print(f"      mesh : {Path(p['mesh_path']).relative_to(results_root)}")
        if p["liveview_png_path"]:
            print(f"      png  : {Path(p['liveview_png_path']).relative_to(data_root)}")
    if skipped:
        print(f"\n[batch] Skipped: {len(skipped)}")
        for name, reason in skipped:
            print(f"  - {name:30s}  {reason}")
    return pairs


# =============================================================================
# Per-sample run (lifted essentially verbatim from legacy batch.py)
# =============================================================================

def _run_single_sample(
    pair: dict,
    cfg: RegistrationConfig,
    aruco_dict_cv: int,
    torch_device: Optional[str],
    use_torch_render: bool,
    render_fn,
) -> dict:
    sample_name = pair["sample_name"]
    output_dir = pair["output_dir"]
    os.makedirs(output_dir, exist_ok=True)
    roi_mode = pair["liveview_png_path"] is not None

    print("\n" + "=" * 68)
    print(f"  BATCH RUN — {sample_name}    ROI={'YES' if roi_mode else 'NO'}")
    print(f"  Output: {output_dir}")
    print("=" * 68)

    t_total = time.time()
    save_pointcloud = cfg.export.save_pointcloud
    save_images = cfg.export.save_images

    res_pc = cfg.render.resolution_mm_per_px
    res_reg = cfg.render.resolution_reg_mm_per_px
    dual = abs(res_reg - res_pc) > 1e-6

    # GPU pre-render once per sample.
    precomputed_render = None
    precomputed_render_reg = None
    elapsed_render = 0.0
    if use_torch_render and render_fn is not None:
        mesh = load_mesh(pair["mesh_path"], scale_m_to_mm=True)
        t0 = time.time()
        precomputed_render = render_fn(
            mesh,
            resolution_mm_per_px=res_pc,
            margin_mm=cfg.render.margin_mm,
            device=torch_device,
        )
        elapsed_render += time.time() - t0
        if dual:
            t0 = time.time()
            precomputed_render_reg = render_fn(
                mesh,
                resolution_mm_per_px=res_reg,
                margin_mm=cfg.render.margin_mm,
                device=torch_device,
            )
            elapsed_render += time.time() - t0
        else:
            precomputed_render_reg = precomputed_render

    t_pipe = time.time()
    result = run_full_pipeline_roi(
        hsi_hdr_path             = pair["hsi_hdr_path"],
        mesh_path                = pair["mesh_path"],
        aruco_json_path          = pair["aruco_path"],
        output_dir               = output_dir,
        liveview_png_path        = pair["liveview_png_path"],
        roi_align_cfg            = cfg.roi.model_dump() if roi_mode else None,
        hsi_extraction_method    = cfg.detection.hsi_extraction_method,
        aruco_dict_type          = aruco_dict_cv,
        suspicious_pixels_hsi    = None,
        render_resolution_mm     = res_pc,
        render_resolution_reg_mm = res_reg,
        render_margin_mm         = cfg.render.margin_mm,
        marker_side_mm           = cfg.detection.marker_side_mm,
        use_subpix               = cfg.detection.use_subpix,
        subpix_winsize           = cfg.detection.subpix_winsize,
        border_px                = cfg.pointcloud.border_px,
        reflectance_norm         = cfg.pointcloud.reflectance_norm,
        pc_chunk_size            = cfg.pointcloud.pc_chunk_size,
        export_ply_file          = cfg.export.ply  if save_pointcloud else False,
        export_npy_file          = cfg.export.npy  if save_pointcloud else False,
        export_csv_file          = cfg.export.csv  if save_pointcloud else False,
        csv_max_points           = cfg.export.csv_max_points,
        save_pointcloud          = save_pointcloud,
        save_images              = save_images,
        sample_name              = sample_name,
        precomputed_render       = precomputed_render,
        precomputed_render_reg   = precomputed_render_reg,
    )
    elapsed_pipeline = time.time() - t_pipe
    elapsed_total = time.time() - t_total

    # -- Extract metrics into a flat row (same schema as legacy batch.py) -----
    err2d_px = result.get("err2d_px")
    err3d_dict = result.get("err3d_dict")
    err3d_dict_pc = result.get("err3d_dict_pc")
    data_hsi = result.get("data_hsi")
    roi_info = result.get("roi_align_info")

    px_per_mm = (
        _px_per_mm_from_data(data_hsi, cfg.detection.marker_side_mm)
        if data_hsi else float("nan")
    )

    if err2d_px is not None:
        err2d = np.asarray(err2d_px, dtype=np.float64)
        m2px, med2px, n2 = _stats(err2d)
        if not np.isnan(px_per_mm) and px_per_mm > 0:
            m2mm, med2mm, _ = _stats(err2d / px_per_mm)
        else:
            m2mm = med2mm = float("nan")
    else:
        m2px = med2px = m2mm = med2mm = float("nan")
        n2 = 0

    if err3d_dict is not None:
        m_rb_bil, md_rb_bil, n_rb_bil = _stats(err3d_dict.get("bilinear"))
        m_rb_bic, md_rb_bic, n_rb_bic = _stats(err3d_dict.get("bicubic"))
    else:
        m_rb_bil = md_rb_bil = m_rb_bic = md_rb_bic = float("nan")
        n_rb_bil = n_rb_bic = 0

    if err3d_dict_pc is not None:
        m_pc_bil, md_pc_bil, n_pc_bil = _stats(err3d_dict_pc.get("bilinear"))
        m_pc_bic, md_pc_bic, n_pc_bic = _stats(err3d_dict_pc.get("bicubic"))
    else:
        m_pc_bil, md_pc_bil = m_rb_bil, md_rb_bil
        m_pc_bic, md_pc_bic = m_rb_bic, md_rb_bic
        n_pc_bil, n_pc_bic = n_rb_bil, n_rb_bic

    return {
        "sample":                    sample_name,
        "roi_mode":                  "yes" if roi_mode else "no",
        "res_reg_mm_pix":            res_reg,
        "res_pc_mm_pix":             res_pc,
        "2D_mean_px":                m2px,
        "2D_median_px":              med2px,
        "2D_mean_mm":                m2mm,
        "2D_median_mm":              med2mm,
        "3D_REG_bilinear_mean_mm":   m_rb_bil,
        "3D_REG_bilinear_median_mm": md_rb_bil,
        "3D_REG_bicubic_mean_mm":    m_rb_bic,
        "3D_REG_bicubic_median_mm":  md_rb_bic,
        "3D_PC_bilinear_mean_mm":    m_pc_bil,
        "3D_PC_bilinear_median_mm":  md_pc_bil,
        "3D_PC_bicubic_mean_mm":     m_pc_bic,
        "3D_PC_bicubic_median_mm":   md_pc_bic,
        "n_corners_2D":              n2,
        "n_points_cloud":            result.get("n_valid", 0),
        "n_bands":                   result.get("bands", 0),
        "roi_inliers":               roi_info["n_inliers"]            if roi_info else 0,
        "roi_matches":               roi_info["n_good_matches"]       if roi_info else 0,
        "roi_reproj_px":             roi_info["reproj_error_mean_px"] if roi_info else float("nan"),
        "elapsed_render_s":          round(elapsed_render,   2),
        "elapsed_pipeline_s":        round(elapsed_pipeline, 2),
        "elapsed_total_s":           round(elapsed_total,    2),
        "status":                    "ok",
    }


def _empty_row(sample: str, res_reg: float, res_pc: float,
               status: str, roi_mode: bool = False) -> dict:
    nan = float("nan")
    return {
        "sample": sample, "roi_mode": "yes" if roi_mode else "no",
        "res_reg_mm_pix": res_reg, "res_pc_mm_pix": res_pc,
        "2D_mean_px": nan, "2D_median_px": nan, "2D_mean_mm": nan, "2D_median_mm": nan,
        "3D_REG_bilinear_mean_mm":   nan, "3D_REG_bilinear_median_mm":   nan,
        "3D_REG_bicubic_mean_mm":    nan, "3D_REG_bicubic_median_mm":    nan,
        "3D_PC_bilinear_mean_mm":    nan, "3D_PC_bilinear_median_mm":    nan,
        "3D_PC_bicubic_mean_mm":     nan, "3D_PC_bicubic_median_mm":     nan,
        "n_corners_2D": 0, "n_points_cloud": 0, "n_bands": 0,
        "roi_inliers": 0, "roi_matches": 0, "roi_reproj_px": nan,
        "elapsed_render_s": 0.0, "elapsed_pipeline_s": 0.0, "elapsed_total_s": 0.0,
        "status": status,
    }


# =============================================================================
# Excel summary (legacy structure)
# =============================================================================

def _write_batch_summary_xlsx(rows: list[dict], output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Summary"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="305496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")
    norm_font = Font(name="Calibri", size=10)

    sample_fill = PatternFill("solid", start_color="FFF2CC")
    roi_fill    = PatternFill("solid", start_color="FCE4D6")
    res_fill    = PatternFill("solid", start_color="F8CBAD")
    err2d_fill  = PatternFill("solid", start_color="DDEBF7")
    reg_fill    = PatternFill("solid", start_color="D9E1F2")
    pc_fill     = PatternFill("solid", start_color="E2EFDA")
    meta_fill   = PatternFill("solid", start_color="F2F2F2")
    err_fill    = PatternFill("solid", start_color="F4B084")

    ws.cell(row=1, column=1, value=f"Batch Summary — {len(rows)} samples").font = \
        Font(bold=True, size=14, color="1F4E78")
    ws.cell(row=2, column=1,
            value=f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}").font = \
        Font(italic=True, color="595959")

    columns: list[tuple[str, str, PatternFill]] = [
        ("sample",                    "Sample",                       sample_fill),
        ("roi_mode",                  "ROI\nmode",                    roi_fill),
        ("res_reg_mm_pix",            "res_reg\n(mm/px)",             res_fill),
        ("res_pc_mm_pix",             "res_pc\n(mm/px)",              res_fill),
        ("2D_mean_px",                "2D mean\n(px)",                err2d_fill),
        ("2D_median_px",              "2D median\n(px)",              err2d_fill),
        ("2D_mean_mm",                "2D mean\n(mm)",                err2d_fill),
        ("2D_median_mm",              "2D median\n(mm)",              err2d_fill),
        ("3D_REG_bilinear_mean_mm",   "3D REG bil\nmean (mm)",        reg_fill),
        ("3D_REG_bilinear_median_mm", "3D REG bil\nmedian (mm)",      reg_fill),
        ("3D_REG_bicubic_mean_mm",    "3D REG bic\nmean (mm)",        reg_fill),
        ("3D_REG_bicubic_median_mm",  "3D REG bic\nmedian (mm)",      reg_fill),
        ("3D_PC_bilinear_mean_mm",    "3D PC bil\nmean (mm)",         pc_fill),
        ("3D_PC_bilinear_median_mm",  "3D PC bil\nmedian (mm)",       pc_fill),
        ("3D_PC_bicubic_mean_mm",     "3D PC bic\nmean (mm)",         pc_fill),
        ("3D_PC_bicubic_median_mm",   "3D PC bic\nmedian (mm)",       pc_fill),
        ("n_corners_2D",              "N corners 2D",                 meta_fill),
        ("n_points_cloud",            "N points",                     meta_fill),
        ("n_bands",                   "N bands",                      meta_fill),
        ("roi_inliers",               "ROI inliers",                  roi_fill),
        ("roi_matches",               "ROI matches",                  roi_fill),
        ("roi_reproj_px",             "ROI reproj (px)",              roi_fill),
        ("elapsed_render_s",          "Render (s)",                   meta_fill),
        ("elapsed_pipeline_s",        "Pipeline (s)",                 meta_fill),
        ("elapsed_total_s",           "Total (s)",                    meta_fill),
        ("status",                    "Status",                       meta_fill),
    ]

    HEADER_ROW = 4
    for ci, (_, label, _) in enumerate(columns, start=1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = border
    ws.row_dimensions[HEADER_ROW].height = 36

    def _fmt(v):
        return "N/A" if isinstance(v, float) and np.isnan(v) else v

    for row_i, row in enumerate(rows, start=HEADER_ROW + 1):
        is_err = row.get("status") != "ok"
        for ci, (key, _, fill) in enumerate(columns, start=1):
            val = row.get(key, "")
            c = ws.cell(row=row_i, column=ci, value=_fmt(val))
            c.alignment = center; c.border = border; c.font = norm_font
            c.fill = err_fill if is_err else fill
            if isinstance(val, float) and not np.isnan(val):
                if key.endswith(("_mm", "_px", "_s")):
                    c.number_format = "0.0000"

    for ci, (_, label, _) in enumerate(columns, start=1):
        max_len = max(len(line) for line in label.split("\n"))
        ws.column_dimensions[get_column_letter(ci)].width = max(max_len + 3, 12)

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=2)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"\n[batch] Aggregate Excel saved → {output_path}")


# =============================================================================
# Entry point
# =============================================================================

def run_batch(
    cfg: RegistrationConfig,
    aruco_dict_cv: int,
    torch_device: Optional[str],
    use_torch_render: bool,
    data_root: Path,
    results_root: Path,
    render_fn=None,
) -> str:
    """Discover every ready sample under ``data_root`` and run them all.

    Per-sample outputs go to ``<results_root>/<sample>/registration/``. The
    aggregate Excel goes to ``<results_root>/_batch/batch_summary.xlsx``.
    """
    roi_mode = bool(cfg.batch.roi_mode)
    print("\n" + "=" * 68)
    print("  BATCH MODE")
    print("=" * 68)
    print(f"  data_root    : {data_root}")
    print(f"  results_root : {results_root}")
    print(f"  ROI mode     : {'YES' if roi_mode else 'NO'}")

    pairs = discover_samples(
        data_root=data_root,
        results_root=results_root,
        roi_mode=roi_mode,
        sample_regex=cfg.batch.sample_regex,
    )
    if not pairs:
        print("\n[batch] No ready samples found. Exiting.")
        return ""

    rows: list[dict] = []
    t_total = time.time()
    for i, pair in enumerate(pairs):
        print(f"\n[batch] Progress: {i + 1}/{len(pairs)} → {pair['sample_name']}")
        try:
            row = _run_single_sample(
                pair=pair,
                cfg=cfg,
                aruco_dict_cv=aruco_dict_cv,
                torch_device=torch_device,
                use_torch_render=use_torch_render,
                render_fn=render_fn,
            )
        except Exception as e:
            print(f"\n[batch] ERROR on {pair['sample_name']}: {e}")
            traceback.print_exc()
            row = _empty_row(
                sample=pair["sample_name"],
                res_reg=cfg.render.resolution_reg_mm_per_px,
                res_pc=cfg.render.resolution_mm_per_px,
                status=f"error: {type(e).__name__}: {e}",
                roi_mode=pair["liveview_png_path"] is not None,
            )
        rows.append(row)

    out_path = str(results_root / "_batch" / "batch_summary.xlsx")
    _write_batch_summary_xlsx(rows, out_path)

    n_ok = sum(1 for r in rows if r["status"] == "ok")
    total_elapsed = time.time() - t_total
    print("\n" + "=" * 68)
    print("BATCH COMPLETED")
    print("=" * 68)
    print(f"  Samples processed : {len(rows)}")
    print(f"  OK / Failed       : {n_ok} / {len(rows) - n_ok}")
    print(f"  Total time        : {total_elapsed:.1f}s ({total_elapsed/60:.1f} min)")
    print(f"  Excel summary     : {out_path}")
    print("=" * 68)
    return out_path


__all__ = ["discover_samples", "run_batch"]
